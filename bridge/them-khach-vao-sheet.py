#!/usr/bin/env python3
"""
them-khach-vao-sheet.py — THÊM KHÁCH VÀO DANH SÁCH GOOGLE SHEET ZALO (an toàn, chỉ append dòng cuối).
===================================================================================================
Diamond: "Tạo danh sách Google sheet để mai thủ công gửi tin nhắn" + "Thu thập email để làm email marketing sau".

CÁCH DÙNG (từ terminal):
  python3 them-khach-vao-sheet.py --ten "Anh Tuấn" --zalo "0912.345.678" \
      --email "tuan@gmail.com" --nhu-cau "Xây mới 5x16, 2 tầng" --nguon "Môi giới: chị Hoa"

Ghi thêm NHIỀU dòng cùng lúc (khuyến cáo: gõ 1 lệnh gom cả buổi khảo sát):
  python3 them-khach-vao-sheet.py --batch danh-sach.csv
    (CSV cột: ten,zalo,email,nhu_cau,nguon — dòng đầu là tiêu đề, bỏ qua)

QUAN TRỌNG (an toàn dữ liệu):
  - Chỉ APPEND dòng vào CUỐI sheet. KHÔNG đụng/ghi đè các dòng khách đã có.
  - KHÔNG xóa sheet, không sửa header.
  - Nên ĐÓNG tab Google Sheet trước khi chạy để tránh xung đột ghi (Google Drive sync local).
  - Sheet mặc định: /mnt/g/My Drive/2026/xaydung/Danh sách khách Zalo.xlsx
"""
import argparse, csv, os, sys
from datetime import datetime
from pathlib import Path

DEFAULT_SHEET = "/mnt/g/My Drive/2026/xaydung/Danh sách khách Zalo.xlsx"

HEADERS = ["STT","Tên khách","Số Zalo / Zalo ID","Email (marketing sau)","Nhu cầu (xây mới/sửa/sơn...)","Nguồn (môi giới nào)","Ngày thêm","Đã gửi Zalo?","Ghi chú"]

def open_wb(path):
    from openpyxl import load_workbook
    if not Path(path).exists():
        print(f"❌ Không thấy sheet: {path}\n   Chạy trước lệnh tạo (hoặc kiểm tra /mnt/g đã mount).")
        sys.exit(1)
    return load_workbook(path)

def append_rows(path, rows):
    wb = open_wb(path)
    ws = wb.active
    # Tìm dòng trống đầu tiên dưới header (hàng 1) — append an toàn
    r = ws.max_row + 1
    # Nếu sheet trống (chưa có header), tự tạo
    if r == 1:
        for c, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
        r = 2
    today = datetime.now().strftime("%d/%m/%Y")
    for row in rows:
        ten, zalo, email, nhucau, nguon = (row + [""]*5)[:5]
        # Ghi vào cột B,C,D,E,F (bỏ cột STT tự điền)
        ws.cell(row=r, column=2, value=(ten or "").strip())
        ws.cell(row=r, column=3, value=(zalo or "").strip())
        ws.cell(row=r, column=4, value=(email or "").strip())
        ws.cell(row=r, column=5, value=(nhucau or "").strip())
        ws.cell(row=r, column=6, value=(nguon or "").strip())
        ws.cell(row=r, column=7, value=today)
        ws.cell(row=r, column=9, value="")  # ghi chú trống
        r += 1
    wb.save(path)
    return len(rows)

def main():
    ap = argparse.ArgumentParser(description="Thêm khách vào Google Sheet Zalo (append cuối, an toàn)")
    ap.add_argument("--ten"); ap.add_argument("--zalo"); ap.add_argument("--email")
    ap.add_argument("--nhu-cau"); ap.add_argument("--nguon")
    ap.add_argument("--batch", help="Đường dẫn CSV nhiều dòng (ten,zalo,email,nhu_cau,nguon)")
    ap.add_argument("--sheet", default=DEFAULT_SHEET, help="Đường dẫn file .xlsx")
    a = ap.parse_args()

    rows = []
    if a.batch:
        if not Path(a.batch).exists():
            print(f"❌ Không thấy file CSV: {a.batch}"); sys.exit(1)
        with open(a.batch, encoding="utf-8") as f:
            rd = csv.reader(f)
            next(rd, None)  # bỏ header
            for line in rd:
                if line and any(c.strip() for c in line):
                    rows.append(line[:5])
    else:
        if not (a.ten or a.zalo or a.email):
            print("⚠️ Cần ít nhất 1 trong: --ten / --zalo / --email. Dùng --help để xem."); sys.exit(1)
        rows.append([a.ten or "", a.zalo or "", a.email or "", a.nhu_cau or "", a.nguon or ""])

    n = append_rows(a.sheet, rows)
    print(f"✅ Đã thêm {n} khách vào sheet:\n   {a.sheet}")
    print("   Mở Google Drive → mở file → thấy khách ở cuối. Đóng sheet trước khi chạy.")

if __name__ == "__main__":
    main()
